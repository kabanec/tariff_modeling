from flask import Flask, request, jsonify, render_template, Response, send_file
from flask_cors import CORS
import random
import os
from dotenv import load_dotenv
from functools import wraps
import requests
from datetime import datetime
import re
import uuid
import logging
import traceback
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG)

API_BASE_URL = os.getenv("API_URL", "https://info.dev.3ceonline.com/ccce/apis")
API_TOKEN = os.getenv("API_TOKEN", "your_token_here")
VALID_USER = os.getenv("AUTH_USER", "admin")
VALID_PASS = os.getenv("AUTH_PASS", "password")


def auth_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        request_id = str(uuid.uuid4())
        auth = request.authorization
        logger.debug(f"[{request_id}] Authorization header: {auth}")
        if not auth:
            logger.error(f"[{request_id}] No authorization header provided")
            return Response('Unauthorized', 401, {'WWW-Authenticate': 'Basic realm="Login Required"'})
        if auth.username != VALID_USER or auth.password != VALID_PASS:
            logger.error(f"[{request_id}] Invalid credentials: username={auth.username}, expected={VALID_USER}")
            return Response('Unauthorized', 401, {'WWW-Authenticate': 'Basic realm="Login Required"'})
        logger.debug(f"[{request_id}] Authentication successful")
        return f(*args, **kwargs)

    return decorated


load_dotenv()

app = Flask(__name__)
CORS(app)

# Avalara API Configuration (Token-based only)
AVALARA_API_BASE = 'https://ns1-quoting-sbx.xbo.avalara.com/api/v2'
AVALARA_TOKEN = os.getenv('AVALARA_TOKEN')
AVALARA_COMPANY_ID = os.getenv('AVALARA_COMPANY_ID')

# ISO-2 Country codes with proper flag emojis (alphabetically sorted)
COUNTRIES = [
    {"code": "AT", "name": "Austria", "flag": "🇦🇹"},
    {"code": "AU", "name": "Australia", "flag": "🇦🇺"},
    {"code": "BE", "name": "Belgium", "flag": "🇧🇪"},
    {"code": "BR", "name": "Brazil", "flag": "🇧🇷"},
    {"code": "CA", "name": "Canada", "flag": "🇨🇦"},
    {"code": "CN", "name": "China", "flag": "🇨🇳"},
    {"code": "CZ", "name": "Czech Republic", "flag": "🇨🇿"},
    {"code": "DE", "name": "Germany", "flag": "🇩🇪"},
    {"code": "DK", "name": "Denmark", "flag": "🇩🇰"},
    {"code": "ES", "name": "Spain", "flag": "🇪🇸"},
    {"code": "FI", "name": "Finland", "flag": "🇫🇮"},
    {"code": "FR", "name": "France", "flag": "🇫🇷"},
    {"code": "GB", "name": "United Kingdom", "flag": "🇬🇧"},
    {"code": "GR", "name": "Greece", "flag": "🇬🇷"},
    {"code": "HK", "name": "Hong Kong", "flag": "🇭🇰"},
    {"code": "ID", "name": "Indonesia", "flag": "🇮🇩"},
    {"code": "IE", "name": "Ireland", "flag": "🇮🇪"},
    {"code": "IN", "name": "India", "flag": "🇮🇳"},
    {"code": "IT", "name": "Italy", "flag": "🇮🇹"},
    {"code": "JP", "name": "Japan", "flag": "🇯🇵"},
    {"code": "KR", "name": "South Korea", "flag": "🇰🇷"},
    {"code": "MX", "name": "Mexico", "flag": "🇲🇽"},
    {"code": "MY", "name": "Malaysia", "flag": "🇲🇾"},
    {"code": "NL", "name": "Netherlands", "flag": "🇳🇱"},
    {"code": "NO", "name": "Norway", "flag": "🇳🇴"},
    {"code": "PH", "name": "Philippines", "flag": "🇵🇭"},
    {"code": "PL", "name": "Poland", "flag": "🇵🇱"},
    {"code": "PT", "name": "Portugal", "flag": "🇵🇹"},
    {"code": "RU", "name": "Russia", "flag": "🇷🇺"},
    {"code": "SE", "name": "Sweden", "flag": "🇸🇪"},
    {"code": "SG", "name": "Singapore", "flag": "🇸🇬"},
    {"code": "TH", "name": "Thailand", "flag": "🇹🇭"},
    {"code": "TR", "name": "Turkey", "flag": "🇹🇷"},
    {"code": "US", "name": "United States", "flag": "🇺🇸"},
    {"code": "VN", "name": "Vietnam", "flag": "🇻🇳"},
    {"code": "Not Specified", "name": "Not Specified", "flag": "🌍"}
]

# Mock HS Code data
HS_CODES = {
    "8517.62.00": {"description": "Machines for reception, conversion of voice, image", "baseline_rate": 0.0},
    "6109.10.00": {"description": "T-shirts, cotton", "baseline_rate": 16.5},
    "8471.30.01": {"description": "Portable computers", "baseline_rate": 0.0},
    "9503.00.00": {"description": "Tricycles, scooters, pedal cars, toys", "baseline_rate": 0.0},
    "8528.72.64": {"description": "Reception apparatus for TV", "baseline_rate": 5.0},
}


def get_avalara_auth_header():
    """Generate Basic Auth header for Avalara API using token"""
    if not AVALARA_TOKEN:
        raise ValueError("Missing AVALARA_TOKEN in environment variables")
    return f"Basic {AVALARA_TOKEN}"


def get_region_for_country(destination_country):
    """Get appropriate region for destination country"""
    region_mapping = {
        'US': 'MA',  # Massachusetts for US
        'CA': 'ON',  # Ontario for Canada
        'MX': 'DF',  # Mexico City for Mexico
    }
    return region_mapping.get(destination_country, '')


def is_valid_hts_code(code):
    """Validate HTS code format"""
    return bool(re.match(r'^\d{4,10}(\.\d{2})?$|^9903\.\d{2}\.\d{2}$|^98\d{2}\.\d{2}\.\d{2}$', code))


def call_global_compliance_api(hs_code, coo, vendor_country, cost_per_unit, quantity, description, import_country,
                               spi_applicable, import_date):
    """
    Call Avalara Global Compliance API for real duty calculations

    Args:
        hs_code: Harmonized System code
        coo: Country of Origin
        vendor_country: Vendor's country (ship from)
        cost_per_unit: Cost per unit in USD
        quantity: Quantity of goods
        description: Product description
        import_country: Import destination country (default US)
        spi_applicable: Whether SPI is applicable (True/False)
        import_date: Import date in YYYY-MM-DD format

    Returns:
        Dictionary with API response and parsed duty data
    """
    try:
        url = f"{AVALARA_API_BASE}/companies/{AVALARA_COMPANY_ID}/globalcompliance"

        # Get auth header
        auth_header = get_avalara_auth_header()

        headers = {
            'Content-Type': 'application/json',
            'Authorization': auth_header
        }

        # Log request details (mask auth for security)
        auth_preview = auth_header[:15] + "..." if len(auth_header) > 15 else auth_header
        logger.debug(f"Making Avalara API call to: {url}")
        logger.debug(f"Auth header preview: {auth_preview}")
        logger.debug(f"Company ID: {AVALARA_COMPANY_ID}")
        logger.debug(f"HS Code: {hs_code}, COO: {coo}, Destination: {import_country}")

        # Build request payload matching exact Avalara format
        payload = {
            "id": "TARIFF-MODEL-001",
            "companyId": int(AVALARA_COMPANY_ID),
            "transactionDate": import_date,
            "currency": "usd",
            "sellerCode": "SELLER-001",
            "b2b": True,
            "shipFrom": {
                "country": vendor_country
            },
            "destinations": [{
                "shipTo": {
                    "country": import_country.lower(),
                    "region": "ca"
                },
                "parameters": [],
                "taxRegistered": False
            }],
            "lines": [{
                "lineNumber": 1,
                "quantity": quantity,
                "preferenceProgramApplicable": spi_applicable,
                "item": {
                    "itemCode": 11,
                    "description": description,
                    "classifications": [{
                        "country": import_country.upper(),
                        "hscode": hs_code.replace(".", "").replace(" ", "")
                    }],
                    "classificationParameters": [{
                        "name": "price",
                        "value": str(cost_per_unit),
                        "unit": "usd"
                    }, {
                        "name": "coo",
                        "value": coo
                    }],
                    "parameters": [{
                        "name": "weight",
                        "value": "0",
                        "unit": "lb"
                    }, {
                        "name": "SHIPPING",
                        "value": "0.00",
                        "unit": "usd"
                    }]
                },
                "classificationParameters": []
            }],
            "type": "QUOTE_MAXIMUM",
            "disableCalculationSummary": False,
            "restrictionsCheck": False,
            "program": "Regular"
        }

        # Make API call
        response = requests.post(url, json=payload, headers=headers, timeout=30)
        response.raise_for_status()

        api_response = response.json()

        # Parse dutyGranularity from response
        duty_lines = []
        total_duty_rate = 0.0

        if 'globalCompliance' in api_response and len(api_response['globalCompliance']) > 0:
            quote = api_response['globalCompliance'][0].get('quote', {})
            lines = quote.get('lines', [])

            if len(lines) > 0:
                calculation_summary = lines[0].get('calculationSummary', {})
                duty_granularity = calculation_summary.get('dutyGranularity', [])

                # Extract each duty type
                for duty in duty_granularity:
                    description_text = duty.get('description', 'Unknown Duty')
                    rate = float(duty.get('rate', 0))
                    rate_percent = rate * 100
                    duty_type = duty.get('type', '')

                    duty_lines.append({
                        'description': description_text,
                        'rate': rate,
                        'rate_percent': rate_percent,
                        'type': duty_type
                    })

                    # Sum up total duty rate
                    total_duty_rate += rate

        # Calculate total duty amount
        total_duty_amount = quantity * cost_per_unit * total_duty_rate

        return {
            'success': True,
            'duty_lines': duty_lines,
            'total_duty_rate': total_duty_rate,
            'total_duty_rate_percent': total_duty_rate * 100,
            'total_duty_amount': total_duty_amount,
            'api_response': api_response,
            'request_payload': payload  # Include request for debugging
        }

    except requests.exceptions.RequestException as e:
        # Capture more detailed error information
        error_details = str(e)
        response_text = None
        status_code = None

        if hasattr(e, 'response') and e.response is not None:
            status_code = e.response.status_code
            try:
                response_text = e.response.json()
            except:
                response_text = e.response.text

            # Specific handling for 401 Unauthorized
            if status_code == 401:
                logger.error("❌ Authentication failed (401 Unauthorized)")
                logger.error(f"URL: {url}")
                logger.error(f"Check your .env file contains:")
                logger.error("  - AVALARA_TOKEN (Base64 encoded)")
                logger.error(f"  - AVALARA_COMPANY_ID={AVALARA_COMPANY_ID}")
                logger.error(f"  - AVALARA_API_BASE={AVALARA_API_BASE}")

                error_details = "Authentication failed. Please verify AVALARA_TOKEN in .env file."

        logger.error(f"API Error [{status_code}]: {error_details}")

        return {
            'success': False,
            'error': error_details,
            'status_code': status_code,
            'error_response': response_text,
            'api_response': None,
            'request_payload': payload if 'payload' in locals() else None
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'api_response': None,
            'request_payload': payload if 'payload' in locals() else None
        }


def calculate_tariff(hs_code, coo, vendor_country, cost_per_unit, quantity, calculation_method="standard"):
    """
    Calculate tariff for a single vendor

    Args:
        hs_code: Harmonized System code
        coo: Country of Origin
        vendor_country: Vendor's country
        cost_per_unit: Cost per unit in USD
        quantity: Quantity of goods
        calculation_method: "standard" or "preferential"

    Returns:
        Dictionary with calculation results
    """

    # Get baseline tariff
    hs_info = HS_CODES.get(hs_code, {"baseline_rate": 0.0})
    baseline_tariff = hs_info.get("baseline_rate", 0.0)

    # Calculate other tariff components
    reciprocal_tariff = 0.0
    chapter_301_tariff = 0.0
    ieepa_tariff = 0.0
    spi_tariff = 0.0

    # Reciprocal tariff logic (example: applies to certain countries)
    if coo in ["CN", "MX", "CA"]:
        reciprocal_tariff = random.uniform(0, 5)

    # Chapter 301 tariff (China specific)
    if coo == "CN":
        chapter_301_tariff = random.uniform(7.5, 25)

    # IEEPA or SPI depending on calculation method
    if calculation_method == "preferential":
        # SPI (Special Preferential Initiative) for preferential calculation
        if coo in ["MX", "CA", "VN"]:
            spi_tariff = random.uniform(0, 10)
    else:
        # IEEPA (International Emergency Economic Powers Act)
        if coo in ["CN", "RU"]:
            ieepa_tariff = random.uniform(10, 25)

    # Calculate total tariff
    total_tariff_rate = baseline_tariff + reciprocal_tariff + chapter_301_tariff

    if calculation_method == "preferential":
        total_tariff_rate += spi_tariff
    else:
        total_tariff_rate += ieepa_tariff

    # Check for de minimis
    customs_value = cost_per_unit * quantity
    duty_deminimis_applied = False

    # US de minimis threshold is $800
    if customs_value < 800:
        duty_deminimis_applied = True
        duty_amount = 0.0
        effective_tariff_rate = 0.0
    else:
        duty_amount = customs_value * (total_tariff_rate / 100)
        effective_tariff_rate = total_tariff_rate

    # Calculate total cost
    total_cost = customs_value + duty_amount

    # Build response
    result = {
        "vendor_country": vendor_country,
        "country_of_origin": coo,
        "cost_per_unit": cost_per_unit,
        "quantity": quantity,
        "customs_value": customs_value,
        "baseline_tariff_rate": baseline_tariff,
        "reciprocal_tariff_rate": reciprocal_tariff,
        "chapter_301_tariff_rate": chapter_301_tariff,
        "total_tariff_rate": effective_tariff_rate,
        "duty_amount": duty_amount,
        "total_cost": total_cost,
        "dutyCalculationSummary": [
            {"name": "DUTY_DEMINIMIS_APPLIED", "value": str(duty_deminimis_applied).lower()}
        ]
    }

    # Add either IEEPA or SPI depending on calculation method
    if calculation_method == "preferential":
        result["spi_tariff_rate"] = spi_tariff
    else:
        result["ieepa_tariff_rate"] = ieepa_tariff

    return result


@app.route('/')
@auth_required
def index():
    """Serve the main application page"""

    incoterms = ['FCA', 'FOB', 'CIF', 'DDP']
    vendors_form = [{'id': i, 'name': '', 'country': '', 'coo': '', 'cost': '', 'quantity': ''} for i in range(1, 7)]
    return render_template('index.html', countries=COUNTRIES, incoterms=incoterms, vendors_form=vendors_form,
                           form_data={})


@app.route('/classify_hs', methods=['POST'])
@auth_required
def classify_hs():
    """
    ADVANCED HS Code Classification using 3CEOnline + Avalara APIs

    This replaces the simple keyword matching with a two-step process:
    1. Call 3CEOnline Classification API to get HS6 code
    2. Call Avalara API with that HS6 code for final classification
    """
    data = request.json
    description = data.get('description', '')

    # Get COO from first vendor or use default
    coo = data.get('coo', 'CN')  # Default to China if not provided
    destination_country = data.get('destination_country', 'US')  # Default to US
    verify_description = data.get('verify_description', False)

    debug_info = f"Request data: {data}\n\n"

    if not description:
        debug_info += "Validation failed: Description missing.\n"
        return jsonify({"error": "Description is required", "debug": debug_info}), 400

    logger.debug(f"Fetching HS code for destination: {destination_country}")

    try:
        # Step 1: Call 3CEOnline classification API to get HS6 code
        classify_url = f"{API_BASE_URL}/classify/v1/interactive/classify-start"
        classify_headers = {
            "Authorization": f"Bearer {API_TOKEN}",
            "Content-Type": "application/json"
        }
        classify_payload = {
            "proddesc": description
        }

        logger.debug(f"Sending 3CEOnline classification request to {classify_url} with payload: {classify_payload}")
        classify_response = requests.post(classify_url, headers=classify_headers, json=classify_payload, timeout=10)

        logger.debug(f"Classification response status: {classify_response.status_code}")
        logger.debug(f"Classification response text: {classify_response.text}")

        if classify_response.status_code != 200:
            debug_info += f"3CEOnline classification API error ({classify_response.status_code}): {classify_response.text}\n"
            return jsonify({"error": f"Classification API error: {classify_response.text}", "debug": debug_info}), 500

        classify_json = classify_response.json()
        debug_info += f"3CEOnline Classification API Response: {classify_json}\n\n"

        # Extract HS6 code from classification response
        hs6_code = None
        requires_interaction = False

        if classify_json.get('data'):
            data_obj = classify_json['data']
            hs6_code_raw = data_obj.get('hsCode', '')

            # Check if there's a current question that needs answering
            current_question = data_obj.get('currentQuestionInteraction')
            if current_question and not hs6_code_raw:
                requires_interaction = True
                debug_info += f"3CEOnline requires additional classification questions. Current question: {current_question.get('name', 'unknown')}\n"

            # Only use the HS code if it's not empty
            if hs6_code_raw and hs6_code_raw.strip():
                hs6_code = str(hs6_code_raw)
                logger.debug(f"Successfully extracted HS6 code from classification: {hs6_code}")

        if not hs6_code:
            debug_info += f"No HS6 code found in 3CEOnline classification response. Raw hsCode value: '{classify_json.get('data', {}).get('hsCode', 'N/A')}'\n"

            # Check if this is because interactive classification is needed
            if requires_interaction:
                debug_info += "Classification requires answering additional questions in 3CEOnline interactive system.\n"

            # If verification is enabled and no HS code found, return verification failure
            if verify_description:
                logger.debug(f"Verification enabled and no HS6 code found for description: '{description}'")
                return jsonify({
                    "verification_failed": True,
                    "error": "Description insufficient for classification - may require more specific details or interactive classification",
                    "debug": debug_info
                }), 200
            else:
                # Verification disabled: proceed to Avalara API with description only (no HS6 code)
                logger.debug(
                    f"No HS6 code from 3CEOnline, but verification disabled. Proceeding to Avalara with description only.")
                debug_info += "Proceeding to Avalara quoting API with description only (no HS6 code from classification).\n"

        # Step 2: Call Avalara API (either with HS6 code from classification, or with description only)
        avalara_headers = {
            "Authorization": get_avalara_auth_header(),
            "Content-Type": "application/json"
        }

        # Build classification parameters - include HS6 code only if we have one
        classification_params = [{"name": "price", "value": "100", "unit": "USD"}]
        if hs6_code:
            classification_params.append({"name": "hs_code", "value": hs6_code})
            debug_info += f"Adding HS6 code {hs6_code} to Avalara request.\n"
        else:
            debug_info += "No HS6 code available - Avalara will classify based on description only.\n"

        # Use actual destination country
        destination_region = get_region_for_country(destination_country)

        # Build Avalara quoting URL
        avalara_url = f"{AVALARA_API_BASE}/companies/{AVALARA_COMPANY_ID}/globalcompliance"

        avalara_payload = {
            "id": "classification-request",
            "companyId": int(AVALARA_COMPANY_ID),
            "currency": "USD",
            "sellerCode": "SC8104341",
            "shipFrom": {"country": coo},
            "destinations": [{"shipTo": {"country": destination_country, "region": destination_region}}],
            "lines": [{
                "lineNumber": 1,
                "quantity": 1,
                "item": {
                    "itemCode": "1",
                    "description": description,
                    "itemGroup": "General",
                    "classificationParameters": classification_params,
                    "parameters": []
                },
                "classificationParameters": classification_params
            }],
            "type": "QUOTE_ENHANCED10",
            "disableCalculationSummary": False,
            "restrictionsCheck": True,
            "program": "Regular"
        }

        logger.debug(
            f"Sending Avalara request to {avalara_url} with destination {destination_country} and payload: {avalara_payload}")
        avalara_response = requests.post(avalara_url, headers=avalara_headers, json=avalara_payload, timeout=10)
        avalara_response.raise_for_status()
        avalara_json = avalara_response.json()
        debug_info += f"Avalara API Response: {avalara_json}\n\n"

        # Extract HS code from Avalara response
        hs_code = avalara_json.get('globalCompliance', [{}])[0].get('quote', {}).get('lines', [{}])[0].get('hsCode')

        if hs_code:
            logger.debug(f"Successfully extracted final HS code from Avalara: {hs_code}")
            return jsonify({
                "hs_code": hs_code,
                "description": f"Classified via 3CEOnline + Avalara",
                "debug": debug_info
            })
        elif hs6_code:
            # Fallback to classified HS6 code if Avalara doesn't provide one
            logger.debug(f"No HS code from Avalara, using classified HS6: {hs6_code}")
            return jsonify({
                "hs_code": hs6_code,
                "description": f"Classified via 3CEOnline (HS6)",
                "debug": debug_info
            })
        else:
            # Neither 3CEOnline nor Avalara provided an HS code
            debug_info += "Neither 3CEOnline classification nor Avalara quoting provided an HS code.\n"
            return jsonify(
                {"error": "No HS code found from either classification or quoting API", "debug": debug_info}), 500

    except requests.RequestException as e:
        debug_info += f"Network error: {str(e)}\nResponse text: {getattr(e.response, 'text', 'No response')}\n"
        logger.error(f"Network error: {str(e)}", exc_info=True)
        return jsonify({"error": f"Network error: {str(e)}", "debug": debug_info}), 500
    except Exception as e:
        debug_info += f"Unexpected error: {str(e)}\nStack trace: {traceback.format_exc()}\n"
        logger.error(f"Unexpected error: {str(e)}", exc_info=True)
        return jsonify({"error": f"Unexpected error: {str(e)}", "debug": debug_info}), 500


@app.route('/calculate_vendor', methods=['POST'])
@auth_required
def calculate_vendor():
    """Calculate tariff for a single vendor using Avalara Global Compliance API"""
    data = request.json

    description = data.get('description', '')
    hs_code = data.get('hs_code', '').replace('.', '')
    coo = data.get('coo')
    vendor_country = coo  # Default ship from = COO
    cost = float(data.get('cost', 0))
    quantity = int(data.get('quantity', 1))
    import_country = data.get('import_country', 'US')
    import_date = data.get('import_date', datetime.now().strftime('%Y-%m-%d'))
    spi_applicable = data.get('spi_applicable', False)

    # Call Avalara Global Compliance API
    result = call_global_compliance_api(
        hs_code=hs_code,
        coo=coo,
        vendor_country=vendor_country,
        cost_per_unit=cost,
        quantity=quantity,
        description=description,
        import_country=import_country,
        spi_applicable=spi_applicable,
        import_date=import_date
    )

    if result['success']:
        # Return structured data for dynamic duty rows
        return jsonify({
            'success': True,
            'duty_lines': result['duty_lines'],
            'total_duty_rate': f"{result['total_duty_rate_percent']:.2f}%",
            'total_duty_amount': result['total_duty_amount'],
            'api_response': result['api_response'],
            'request_payload': result.get('request_payload')
        })
    else:
        # Return error with API response for debug
        return jsonify({
            'success': False,
            'error': result['error'],
            'error_response': result.get('error_response'),
            'api_response': result.get('api_response'),
            'request_payload': result.get('request_payload')
        }), 500


@app.route('/export_excel', methods=['POST'])
@auth_required
def export_excel():
    """Export calculation results to Excel"""
    try:
        data = request.json
        form_data = data.get('formData', {})
        vendors = data.get('vendors', [])

        if not vendors:
            return jsonify({'error': 'No vendor data to export'}), 400

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tariff Calculations"

        # Define styles
        header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        subheader_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        total_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = 'Avalara Tariff Modeling - Calculation Results'
        ws['A1'].font = Font(bold=True, size=16, color="FF6600")
        ws['A1'].alignment = center_align

        # Product Information Section
        row = 3
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'Product & Import Details'
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = left_align

        row += 1
        product_info = [
            ('Ship Date:', form_data.get('import_date', 'N/A')),
            ('Destination:', form_data.get('import_country', 'N/A')),
            ('Part #/SKU:', form_data.get('part_sku', 'N/A')),
            ('Product Description:', form_data.get('description', 'N/A')),
            ('Tariff Code (HS Code):', form_data.get('hs_code', 'N/A')),
            ('Order Quantity:', form_data.get('order_qty', 'N/A')),
            ('SPI Applicable:', 'Yes' if form_data.get('spi_applicable') else 'No')
        ]

        for label, value in product_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:F{row}')
            row += 1

        # Vendor Comparison Section
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'Vendor Comparison'
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = left_align

        # Get all unique duty types across all vendors
        all_duty_types = set()
        for vendor in vendors:
            for duty_line in vendor.get('duty_lines', []):
                all_duty_types.add(duty_line['description'])
        all_duty_types = sorted(list(all_duty_types))

        # Vendor table headers
        row += 1
        headers_start_row = row
        headers = ['Metric'] + [v.get('name', f"Vendor {i + 1}") for i, v in enumerate(vendors)]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = center_align
            cell.border = border

        # Vendor basic info
        row += 1
        vendor_info_rows = [
            ('Vendor Country', lambda v: v.get('vendor_country', 'N/A')),
            ('Country of Origin', lambda v: v.get('coo', 'N/A')),
            ('COGS per Unit', lambda v: f"${v.get('cost', 0):.2f}"),
            ('Quantity', lambda v: str(v.get('quantity', 0)))
        ]

        for label, value_func in vendor_info_rows:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).alignment = left_align
            ws.cell(row=row, column=1).border = border

            for col_idx, vendor in enumerate(vendors, start=2):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = value_func(vendor)
                cell.alignment = center_align
                cell.border = border
            row += 1

        # Duty breakdown section header
        ws.cell(row=row, column=1).value = 'Duty Breakdown'
        ws.cell(row=row, column=1).font = subheader_font
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        ws.cell(row=row, column=1).border = border
        for col_idx in range(2, len(vendors) + 2):
            ws.cell(row=row, column=col_idx).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF",
                                                                fill_type="solid")
            ws.cell(row=row, column=col_idx).border = border
        row += 1

        # Individual duty lines
        for duty_type in all_duty_types:
            ws.cell(row=row, column=1).value = duty_type
            ws.cell(row=row, column=1).alignment = left_align
            ws.cell(row=row, column=1).border = border

            for col_idx, vendor in enumerate(vendors, start=2):
                # Find matching duty line
                duty_value = 'N/A'
                for duty_line in vendor.get('duty_lines', []):
                    if duty_line['description'] == duty_type:
                        duty_value = f"{duty_line['rate_percent']:.2f}%"
                        break

                cell = ws.cell(row=row, column=col_idx)
                cell.value = duty_value
                cell.alignment = center_align
                cell.border = border
            row += 1

        # Total Duty Rate
        ws.cell(row=row, column=1).value = 'Total Duty Rate'
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=1).border = border

        for col_idx, vendor in enumerate(vendors, start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = vendor.get('total_duty_rate', 'N/A')
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = center_align
            cell.border = border
        row += 1

        # Total Duty Amount
        ws.cell(row=row, column=1).value = 'Total Duty Amount'
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=1).border = border

        for col_idx, vendor in enumerate(vendors, start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = f"${vendor.get('total_duty_amount', 0):.2f}"
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = center_align
            cell.border = border
        row += 1

        # Total Landed Cost
        ws.cell(row=row, column=1).value = 'Total Landed Cost'
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FF6600")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
        ws.cell(row=row, column=1).alignment = left_align
        ws.cell(row=row, column=1).border = border

        for col_idx, vendor in enumerate(vendors, start=2):
            cell = ws.cell(row=row, column=col_idx)
            merchandise_value = vendor.get('cost', 0) * vendor.get('quantity', 0)
            total_landed = merchandise_value + vendor.get('total_duty_amount', 0)
            cell.value = f"${total_landed:.2f}"
            cell.font = Font(bold=True, size=12, color="FF6600")
            cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
            cell.alignment = center_align
            cell.border = border

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        for col_idx in range(2, len(vendors) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        filename = f'Tariff_Calculations_{timestamp}.xlsx'

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/countries', methods=['GET'])
@auth_required
def get_countries():
    """Return list of country codes"""
    return jsonify({"countries": COUNTRIES})


@app.route('/api/hs-codes', methods=['GET'])
@auth_required
def get_hs_codes():
    """Return list of HS codes"""
    return jsonify({"hs_codes": list(HS_CODES.keys())})


@app.route('/api/hs-code/<code>', methods=['GET'])
@auth_required
def get_hs_code_info(code):
    """Return info for specific HS code"""
    info = HS_CODES.get(code)
    if info:
        return jsonify({"hs_code": code, **info})
    else:
        return jsonify({"error": "HS Code not found"}), 404


@app.route('/api/calculate', methods=['POST'])
@auth_required
def calculate():
    """
    Calculate tariff for a single vendor

    Expected JSON body:
    {
        "hs_code": "8517.62.00",
        "country_of_origin": "CN",
        "vendor_country": "CN",
        "cost_per_unit": 100.50,
        "quantity": 1000,
        "calculation_method": "standard"  // or "preferential"
    }
    """
    try:
        data = request.json

        # Validate required fields
        required_fields = ['hs_code', 'country_of_origin', 'cost_per_unit', 'quantity']
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"Missing required field: {field}"}), 400

        # Validate COO is not "Not Specified"
        if data['country_of_origin'] == 'Not Specified':
            return jsonify({"error": "Country of Origin cannot be 'Not Specified' for calculations"}), 400

        # Validate cost > 0
        if data['cost_per_unit'] <= 0:
            return jsonify({"error": "Cost per unit must be greater than 0"}), 400

        # Validate quantity > 0
        if data['quantity'] <= 0:
            return jsonify({"error": "Quantity must be greater than 0"}), 400

        # Get calculation method (default to standard)
        calculation_method = data.get('calculation_method', 'standard')

        # Perform calculation
        result = calculate_tariff(
            hs_code=data['hs_code'],
            coo=data['country_of_origin'],
            vendor_country=data.get('vendor_country', data['country_of_origin']),
            cost_per_unit=data['cost_per_unit'],
            quantity=data['quantity'],
            calculation_method=calculation_method
        )

        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/health', methods=['GET'])
@auth_required
def health():
    """Health check endpoint"""
    return jsonify({"status": "ok"})


if __name__ == '__main__':
    app.run(debug=True, port=5000)